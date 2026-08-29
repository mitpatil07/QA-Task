"""
generate_manual_tests.py
========================
Generates manual-test-cases.xlsx with 4 module sheets + Bug Report.

Run:  python generate_manual_tests.py
Output: ../manual-testing/manual-test-cases.xlsx

Requires: pip install openpyxl
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue
PASS_FILL     = PatternFill("solid", fgColor="C6EFCE")   # light green
FAIL_FILL     = PatternFill("solid", fgColor="FFC7CE")   # light red
BLOCKED_FILL  = PatternFill("solid", fgColor="FFEB9C")   # light yellow
ALT_ROW_FILL  = PatternFill("solid", fgColor="EBF3FB")   # very light blue

HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
NORMAL_FONT  = Font(name="Calibri", size=10)
BOLD_FONT    = Font(name="Calibri", bold=True, size=10)

THIN = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    "Test Case ID", "Module", "Test Scenario", "Precondition",
    "Test Steps", "Test Data", "Expected Result",
    "Actual Result", "Status", "Priority"
]

COL_WIDTHS = [14, 18, 40, 30, 60, 35, 50, 30, 10, 10]


def style_header_row(ws):
    for col_idx, (heading, width) in enumerate(zip(COLUMNS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=heading)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30


def add_row(ws, row_data, row_num):
    fill = ALT_ROW_FILL if row_num % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    status = row_data[8] if len(row_data) > 8 else ""
    if status == "Pass":   fill = PASS_FILL
    elif status == "Fail": fill = FAIL_FILL
    elif status == "Blocked": fill = BLOCKED_FILL

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font      = NORMAL_FONT
        cell.fill      = fill
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row_num].height = 50


def create_sheet(wb, sheet_name, rows):
    ws = wb.create_sheet(title=sheet_name)
    ws.freeze_panes = "A2"
    style_header_row(ws)
    for i, row in enumerate(rows, start=2):
        # Pad row to 10 columns
        padded = list(row) + [""] * (10 - len(row))
        add_row(ws, padded, i)
    return ws


# ---------------------------------------------------------------------------
# Test Case Data
# ---------------------------------------------------------------------------

SIGNUP_CASES = [
    # ID, Module, Scenario, Precondition, Steps, TestData, Expected, Actual, Status, Priority
    ("TC_SU_001", "Sign Up", "Valid sign-up with all required fields",
     "User not registered",
     "1. Go to sign-up page\n2. Enter valid Name, Email, Phone, Password, Confirm Password\n3. Click Sign Up",
     "Name:QA Tester\nEmail:qa@test.com\nPhone:9876543210\nPass:Test@1234",
     "Account created, redirected to login or dashboard", "", "Pass", "High"),

    ("TC_SU_002", "Sign Up", "Sign-up with valid mobile number as identifier",
     "User not registered",
     "1. Go to sign-up page\n2. Enter mobile number instead of email\n3. Fill other fields\n4. Submit",
     "Phone:9876543210\nPass:Test@1234",
     "Account created successfully", "", "", "High"),

    ("TC_SU_003", "Sign Up", "Submit with Name field empty",
     "Sign-up page open",
     "1. Leave Name blank\n2. Fill all other fields\n3. Click Sign Up",
     "Name: (empty)\nEmail:qa@test.com",
     "Error: 'Name is required' shown below field", "", "", "High"),

    ("TC_SU_004", "Sign Up", "Invalid email format",
     "Sign-up page open",
     "1. Enter 'notanemail' in email field\n2. Fill other fields\n3. Click Sign Up",
     "Email: notanemail",
     "Error: 'Invalid email format'", "", "", "High"),

    ("TC_SU_005", "Sign Up", "Password shorter than minimum (weak password)",
     "Sign-up page open",
     "1. Enter password with fewer than 8 characters\n2. Submit",
     "Password: abc",
     "Error: 'Password must be at least 8 characters'", "", "", "High"),

    ("TC_SU_006", "Sign Up", "Mismatched Confirm Password",
     "Sign-up page open",
     "1. Enter Password: Test@1234\n2. Enter Confirm Password: Different@999\n3. Submit",
     "Pass:Test@1234\nConfirm:Different@999",
     "Error: 'Passwords do not match'", "", "", "High"),

    ("TC_SU_007", "Sign Up", "Duplicate email registration",
     "Email already registered",
     "1. Enter already-registered email\n2. Fill other fields\n3. Submit",
     "Email:mitesh8767@gmail.com",
     "Error: 'Email already in use' or similar", "", "", "High"),

    ("TC_SU_008", "Sign Up", "Name field — special characters",
     "Sign-up page open",
     "1. Enter name with special chars: 'Test@#$%'\n2. Submit",
     "Name: Test@#$%",
     "Either accepted or error message shown (document actual behaviour)", "", "", "Medium"),

    ("TC_SU_009", "Sign Up", "Email field — character limit",
     "Sign-up page open",
     "1. Enter 256-character email\n2. Submit",
     "Email: (256 chars)",
     "Field rejects input beyond limit OR validation error shown", "", "", "Low"),

    ("TC_SU_010", "Sign Up", "Phone field — non-numeric characters",
     "Sign-up page open",
     "1. Enter 'abc123' in phone field\n2. Submit",
     "Phone: abc123",
     "Error: 'Enter a valid phone number'", "", "", "Medium"),

    ("TC_SU_011", "Sign Up", "UI — Password masking and eye icon toggle",
     "Sign-up page open",
     "1. Observe password field is masked by default\n2. Click the eye icon\n3. Observe password is revealed\n4. Click again to mask",
     "Password: Test@1234",
     "Password masked by default; eye icon toggles visibility correctly", "", "", "Medium"),

    ("TC_SU_012", "Sign Up", "UI — Required field indicators (asterisks)",
     "Sign-up page open",
     "1. Observe all required fields\n2. Note which have asterisk (*) markers",
     "N/A",
     "All mandatory fields marked with * indicator", "", "", "Low"),
]

FORGOT_PASSWORD_CASES = [
    ("TC_FP_001", "Forgot Password", "Submit valid registered email",
     "User account exists",
     "1. Click 'Forgot Password?' on login page\n2. Enter registered email\n3. Click Submit",
     "Email:mitesh8767@gmail.com",
     "OTP sent confirmation shown (or email dispatched)", "", "", "High"),

    ("TC_FP_002", "Forgot Password", "Enter correct OTP and reset password",
     "OTP received",
     "1. Enter OTP received on email/phone\n2. Enter new password\n3. Confirm new password\n4. Submit",
     "OTP:(valid)\nNewPass:NewPass@123",
     "Password reset success, redirected to login", "", "", "High"),

    ("TC_FP_003", "Forgot Password", "Submit unregistered email",
     "Forgot Password page open",
     "1. Enter email not in the system\n2. Click Submit",
     "Email:notregistered@test.com",
     "Error: 'User not found' or 'Invalid email'", "", "", "High"),

    ("TC_FP_004", "Forgot Password", "Invalid email format",
     "Forgot Password page open",
     "1. Enter 'notanemail' in the field\n2. Click Submit",
     "Email: notanemail",
     "Error: 'Invalid email format'", "", "", "High"),

    ("TC_FP_005", "Forgot Password", "Enter wrong OTP",
     "OTP received, OTP page open",
     "1. Enter incorrect OTP (e.g., 000000)\n2. Submit",
     "OTP: 000000",
     "Error: 'Invalid OTP' or 'Incorrect OTP'", "", "", "High"),

    ("TC_FP_006", "Forgot Password", "OTP expiry — submit after timeout",
     "OTP received, wait for expiry (typically 5-10 min)",
     "1. Wait for OTP to expire\n2. Enter (now expired) OTP\n3. Submit",
     "OTP:(expired)",
     "Error: 'OTP expired, please request a new one'", "", "", "High"),

    ("TC_FP_007", "Forgot Password", "Empty email field — submit",
     "Forgot Password page open",
     "1. Leave email field empty\n2. Click Submit",
     "Email: (empty)",
     "Validation error: 'Email is required'", "", "", "High"),

    ("TC_FP_008", "Forgot Password", "Field-level: OTP accepts only numeric input",
     "OTP entry screen open",
     "1. Try typing alphabets in OTP field\n2. Observe behaviour",
     "OTP input: 'abcdef'",
     "Non-numeric characters rejected or field type=number enforces digits only", "", "", "Medium"),

    ("TC_FP_009", "Forgot Password", "Field-level: OTP field character limit",
     "OTP entry screen open",
     "1. Attempt to type more than 6 digits\n2. Observe",
     "OTP: 1234567 (7 digits)",
     "Field limits input to 4 or 6 characters (as designed)", "", "", "Low"),

    ("TC_FP_010", "Forgot Password", "UI — Resend OTP cooldown timer",
     "OTP page open, OTP just sent",
     "1. Note the 'Resend OTP' button state\n2. Observe countdown timer\n3. Try clicking Resend during cooldown",
     "N/A",
     "Resend button disabled with countdown (e.g., 'Resend in 60s'); enabled after timer", "", "", "Medium"),

    ("TC_FP_011", "Forgot Password", "UI — Back button returns to login page",
     "Forgot Password page open",
     "1. Click 'Back' button",
     "N/A",
     "Navigated back to login page (/auth/login)", "", "Pass", "Low"),
]

OTP_CASES = [
    ("TC_OTP_001", "Sign in with OTP", "Valid mobile — OTP → Login",
     "Registered account with mobile number",
     "1. On login page, select 'Sign in with OTP'\n2. Enter registered mobile\n3. Submit\n4. Enter received OTP\n5. Submit",
     "Mobile: 9876543210\nOTP:(valid)",
     "Logged in, redirected to dashboard", "", "", "High"),

    ("TC_OTP_002", "Sign in with OTP", "Resend OTP — new OTP works",
     "Mobile submitted, OTP page open",
     "1. Click 'Resend OTP' after cooldown\n2. Enter newly received OTP",
     "OTP:(new)",
     "Login successful with new OTP", "", "", "High"),

    ("TC_OTP_003", "Sign in with OTP", "Unregistered mobile number",
     "OTP login page",
     "1. Enter mobile number not registered\n2. Submit",
     "Mobile: 9000000000",
     "Error: 'Mobile number not registered'", "", "", "High"),

    ("TC_OTP_004", "Sign in with OTP", "Wrong OTP entered",
     "OTP received, OTP entry page",
     "1. Enter wrong OTP\n2. Submit",
     "OTP: 000000",
     "Error: 'Invalid OTP'", "", "", "High"),

    ("TC_OTP_005", "Sign in with OTP", "Expired OTP submitted",
     "OTP received, wait for expiry",
     "1. Let OTP expire\n2. Enter expired OTP",
     "OTP:(expired)",
     "Error: 'OTP has expired'", "", "", "High"),

    ("TC_OTP_006", "Sign in with OTP", "Empty OTP field — submit",
     "OTP entry page open",
     "1. Leave OTP field empty\n2. Click Submit",
     "OTP: (empty)",
     "Validation error: 'OTP is required'", "", "", "High"),

    ("TC_OTP_007", "Sign in with OTP", "Field-level: OTP accepts only 6 digits",
     "OTP entry page",
     "1. Try entering 7 digits\n2. Try entering letters",
     "OTP: 1234567 / abcdef",
     "Field limited to 6 numeric digits only", "", "", "Medium"),

    ("TC_OTP_008", "Sign in with OTP", "Resend button disabled during cooldown",
     "OTP just sent",
     "1. Immediately click Resend\n2. Observe button state",
     "N/A",
     "Button disabled; countdown timer shown (e.g. 60 seconds)", "", "", "Medium"),

    ("TC_OTP_009", "Sign in with OTP", "UI — Countdown timer for OTP expiry",
     "OTP entry page",
     "1. Note timer visible on screen\n2. Observe countdown",
     "N/A",
     "Timer displayed and counts down; message shown when expired", "", "", "Medium"),

    ("TC_OTP_010", "Sign in with OTP", "UI — Auto-submit on entering last OTP digit",
     "OTP entry page (if OTP has individual digit boxes)",
     "1. Enter all OTP digits one by one\n2. Observe after last digit",
     "OTP: 123456",
     "Form auto-submits OR Submit button enabled only after all digits filled", "", "", "Low"),
]

LOGIN_CASES = [
    ("TC_L_001", "Login", "Valid login with email and password",
     "Registered account exists",
     "1. Go to /auth/login\n2. Enter valid email + password\n3. Click Login",
     "Email:mitesh8767@gmail.com\nPass:YOUR_PASSWORD",
     "Redirected to dashboard (/dashboard)", "", "Pass", "High"),

    ("TC_L_002", "Login", "Valid login with mobile number and password",
     "Account registered with mobile",
     "1. Enter mobile number in identifier field\n2. Enter password\n3. Click Login",
     "Mobile:9876543210\nPass:YOUR_PASSWORD",
     "Redirected to dashboard", "", "", "High"),

    ("TC_L_003", "Login", "Session persists after page refresh",
     "User logged in",
     "1. Log in successfully\n2. Press F5 to refresh\n3. Observe page",
     "N/A",
     "User remains logged in, dashboard shown", "", "", "Medium"),

    ("TC_L_004", "Login", "Wrong password for registered email",
     "Account exists",
     "1. Enter correct email\n2. Enter wrong password\n3. Click Login",
     "Email:mitesh8767@gmail.com\nPass:WrongPass",
     "'Invalid Email Id / Mobile No or Password'", "", "Pass", "High"),

    ("TC_L_005", "Login", "Unregistered email",
     "Login page open",
     "1. Enter email not in system\n2. Enter any password\n3. Click Login",
     "Email:nobody@test.com\nPass:Test@1234",
     "Error: 'Invalid Email Id / Mobile No or Password'", "", "", "High"),

    ("TC_L_006", "Login", "Empty email/identifier field",
     "Login page open",
     "1. Leave email field empty\n2. Enter password\n3. Click Login",
     "Email:(empty)\nPass:Test@1234",
     "Validation error shown below email field", "", "", "High"),

    ("TC_L_007", "Login", "Empty password field",
     "Login page open",
     "1. Enter valid email\n2. Leave password empty\n3. Click Login",
     "Email:mitesh8767@gmail.com\nPass:(empty)",
     "Validation error shown below password field", "", "", "High"),

    ("TC_L_008", "Login", "Both fields empty",
     "Login page open",
     "1. Leave both fields empty\n2. Click Login",
     "Email:(empty)\nPass:(empty)",
     "Validation errors on both fields", "", "", "High"),

    ("TC_L_009", "Login", "Field-level: email field requires '@' symbol",
     "Login page open",
     "1. Enter 'notanemail' (no @)\n2. Click Login",
     "Email: notanemail",
     "HTML5 or custom validation: 'Please enter a valid email address'", "", "", "Medium"),

    ("TC_L_010", "Login", "Field-level: password minimum length",
     "Login page open",
     "1. Enter password with only 2 characters\n2. Submit",
     "Pass: ab",
     "Error: 'Password too short' or form prevents submission", "", "", "Medium"),

    ("TC_L_011", "Login", "Field-level: email field max character limit",
     "Login page open",
     "1. Paste 300-character string in email field\n2. Observe",
     "Email:(300 chars)",
     "Field truncates input at limit OR accepts all (document behaviour)", "", "", "Low"),

    ("TC_L_012", "Login", "UI — Password masking toggle (eye icon)",
     "Login page open",
     "1. Type password — verify it shows dots/asterisks\n2. Click eye icon — verify plain text shown\n3. Click again — verify masked again",
     "Pass: Test@1234",
     "Password masked by default; eye icon correctly toggles visibility", "", "", "Medium"),

    ("TC_L_013", "Login", "UI — Error message clarity and placement",
     "After submitting invalid credentials",
     "1. Submit wrong credentials\n2. Observe error message location, text, colour",
     "Email:wrong@test.com\nPass:wrong",
     "Error message is prominent, clearly worded, and visible near the form", "", "", "Medium"),

    ("TC_L_014", "Login", "UI — Login button state",
     "Login page open",
     "1. Observe button when fields are empty\n2. Fill fields and observe",
     "N/A",
     "Button may be disabled when fields empty; enabled once filled (document actual)", "", "", "Low"),
]

BUG_REPORT_COLUMNS = [
    "Bug ID", "Module", "Bug Title", "Steps to Reproduce",
    "Expected Result", "Actual Result", "Severity", "Status", "Screenshot Reference"
]

BUG_WIDTHS = [10, 15, 35, 60, 40, 40, 12, 12, 25]

BUG_ROWS = [
    ("UX_001", "Forgot Password",
     "UX Note: Unclear error message when submitting unregistered email",
     "1. Go to /auth/login\n2. Click 'Forgot Password?'\n3. Enter an unregistered email (e.g., unknown@test.com)\n4. Click Submit",
     "Should show a clear, user-friendly message (e.g., 'Email not found')",
     "Displays a generic or vague error message instead of a specific user-friendly prompt",
     "Low (UX)", "Open", "N/A"),
]


# ---------------------------------------------------------------------------
# Build the workbook
# ---------------------------------------------------------------------------

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # ── Module sheets ──────────────────────────────────────────────────────
    create_sheet(wb, "Sign Up",            SIGNUP_CASES)
    create_sheet(wb, "Forgot Password",    FORGOT_PASSWORD_CASES)
    create_sheet(wb, "Sign in with OTP",   OTP_CASES)
    create_sheet(wb, "Login",              LOGIN_CASES)

    # ── Bug Report sheet ───────────────────────────────────────────────────
    ws_bug = wb.create_sheet(title="Bug Report")
    ws_bug.freeze_panes = "A2"

    for col_idx, (heading, width) in enumerate(zip(BUG_REPORT_COLUMNS, BUG_WIDTHS), start=1):
        cell = ws_bug.cell(row=1, column=col_idx, value=heading)
        cell.font      = HEADER_FONT
        cell.fill      = PatternFill("solid", fgColor="7B2C2C")  # dark red header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws_bug.column_dimensions[get_column_letter(col_idx)].width = width
    ws_bug.row_dimensions[1].height = 30

    for i, row in enumerate(BUG_ROWS, start=2):
        padded = list(row) + [""] * (9 - len(row))
        sev = padded[6]
        fill = (PatternFill("solid", fgColor="FFC7CE") if sev == "High"
                else PatternFill("solid", fgColor="FFEB9C") if sev == "Medium"
                else PatternFill("solid", fgColor="C6EFCE"))
        for col_idx, value in enumerate(padded, start=1):
            cell = ws_bug.cell(row=i, column=col_idx, value=value)
            cell.font      = NORMAL_FONT
            cell.fill      = fill
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws_bug.row_dimensions[i].height = 60

    # ── Save ───────────────────────────────────────────────────────────────
    output_dir  = os.path.join(os.path.dirname(__file__), "..", "manual-testing")
    output_path = os.path.join(output_dir, "manual-test-cases.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(output_path)
    print(f"[DONE] Saved: {os.path.abspath(output_path)}")
    print(f"    Sheets: Sign Up, Forgot Password, Sign in with OTP, Login, Bug Report")


if __name__ == "__main__":
    main()
